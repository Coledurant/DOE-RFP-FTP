from bs4 import BeautifulSoup
import requests
import warnings
import urllib.request
from urllib.request import urlopen
from contextlib import closing
import shutil
import re
from collections import Counter
import os
import json
import sys
from datetime import datetime, timedelta
import logging
from send_email import send_email
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import operator
warnings.filterwarnings("ignore", category=UserWarning, module='bs4')

logger = logging.getLogger(__name__)

try:
    import configparser
except ImportError:
    import ConfigParser as configparser

from classes import ConEdisonRFP, ConEdisonDocument

###############################################################################
###############################################################################
###############################################################################

conf = configparser.ConfigParser()
config_file = os.path.join(os.path.dirname(__file__), "config.ini")
conf.read(config_file)
phrases_config = conf.get('all', 'phrases')
phrases = phrases_config.split(',')

###############################################################################
###############################################################################
###############################################################################

# All
###############################################################################

def check_if_new_file(file_name):

    files = os.listdir()

    if file_name in files:
        new=False
    else:
        new=True

    assert isinstance(new, bool), 'Download check did not return a boolean'

    return new

def download_pdf(download_url, document_name):

    '''
    Downloads a pdf at the download_url location and saves it as document_name
    as lonf as document_name ends if .pdf
    Parameters:
        download_url (str): URL link for the pdf to download
        document_name (str): Name to save the downloaded pdf under (must end in .pdf)
    Returns:
        None - Just downloads the pdf
    '''

    assert document_name.endswith('.pdf'), 'Document name did not end with .pdf'

    if check_if_new_file(document_name):
        response = urlopen(download_url)

        file = open(document_name, 'wb')
        file.write(response.read())
        file.close()

        inf = os.getcwd().split('RFPFinder' + os.sep + 'data')
        file_path = inf[1]

        history('pdf_download', pdf_name = document_name, file_path = file_path)
    else:pass

    return

def download_excel(download_url, document_name):

    assert document_name.endswith('.xlsx'), 'Document name did not end with .xlsx'

    if check_if_new_file(document_name):

        urllib.request.urlretrieve(download_url, document_name)

        inf = os.getcwd().split('RFPFinder' + os.sep + 'data')
        file_path = inf[1]

        history('excel_download', excel_name = document_name, file_path = file_path)
    else:pass

def history(change_type='run', **kwargs):

    '''
    Goal is to try to keep a file of all new downloads and when they occur
    '''
    begin_dir = os.getcwd()

    now = datetime.now()
    hist_date = now.strftime("%m/%d/%Y")
    hist_time = now.strftime("%H:%M")

    os.chdir(data_dir)

    book = load_workbook('history.xlsx')

    if change_type == 'run':

        sheet = book['Runs']

        # Time, Run
        sheet.append([now, 'Run'])

    elif change_type == 'pdf_download':

        pdf_name = kwargs.get('pdf_name')
        file_path = kwargs.get('file_path')

        sheet = book['PDF Downloads']

        # Time, Change Type, PDF Name, Location
        sheet.append([now, 'PDF Download', pdf_name, file_path])

    elif change_type == 'fbo_daily_message':

        hasdata = kwargs.get('hasdata')

        sheet = book['FBO FTP Access']
        now_minus_two = datetime.utcnow() - timedelta(2)

        # Time, FTP Access, hasdata
        sheet.append([now, now_minus_two, 'FTP Access', hasdata])

    elif change_type == 'ny_rev_connect':

        ny_rev_area = kwargs.get('ny_rev_area')
        ny_rev_info = kwargs.get('ny_rev_info')

        sheet = book['NY Rev Connect']

        # Time, NY Rev Area, Info
        sheet.append([now, ny_rev_area, ny_rev_info])

    elif change_type == 'created_dir':

        dir_location = kwargs.get('dir_location')

        sheet = book['Created Folder']

        sheet.append([now, dir_location])

    elif change_type == 'aep':

        sheet = book['AEP']

        aep_type = kwargs.get('aep_type')
        aep_area = kwargs.get('aep_area')
        aep_info = kwargs.get('aep_info')


        # Now, Change Type, Area Location, Information
        sheet.append([now, aep_type, aep_area, aep_info])

    elif change_type == 'excel_download':

        excel_name = kwargs.get('excel_name')
        file_path = kwargs.get('file_path')

        sheet = book['Excel Downloads']

        # Time, Change Type, PDF Name, Location
        sheet.append([now, 'Excel Download', excel_name, file_path])

    book.save('history.xlsx')
    os.chdir(begin_dir)
    return None

###############################################################################
###############################################################################
###############################################################################

# ROOT DIRS
curr_dir = os.getcwd()
data_dir = os.path.join(curr_dir, 'data')
if not os.path.exists(data_dir):
    os.mkdir(data_dir)
    history('created_dir', dir_location = data_dir.split('RFPFinder')[1])
else:pass

# AEP DIRS
aep_dir = os.path.join(data_dir, 'AEP')
if not os.path.exists(aep_dir):
    os.mkdir(aep_dir)
    history('created_dir', dir_location = aep_dir.split('RFPFinder')[1])
else:pass
aep_ohio_dir = os.path.join(aep_dir, 'AEP Ohio')
aep_texas_dir = os.path.join(aep_dir, 'AEP Texas')
appalachian_power_dir = os.path.join(aep_dir, 'Appalachian Power')
indiana_michigan_dir = os.path.join(aep_dir, 'Indiana Michigan')
kentucky_power_dir = os.path.join(aep_dir, 'Kentucky Power')
public_service_company_of_oklahoma_dir = os.path.join(aep_dir, 'Public Service Company of Oklahoma')
southwestern_electric_power_company_dir = os.path.join(aep_dir, 'Southwestern Electric Power Company')

# Puerto Rico Government DIRS
puerto_rico_government_dir = os.path.join(data_dir, 'Puerto Rico Government')
if not os.path.exists(puerto_rico_government_dir):
    os.mkdir(puerto_rico_government_dir)
    history('created_dir', dir_location = puerto_rico_government_dir.split('RFPFinder')[1])
else:pass
puerto_rico_government_pdf_dir = os.path.join(puerto_rico_government_dir, 'puerto_rico_government_pdfs')

# NY Rev Connect DIRS
ny_rev_connect_dir = os.path.join(data_dir, 'NY Rev Connect')
if not os.path.exists(ny_rev_connect_dir):
    os.mkdir(ny_rev_connect_dir)
    history('created_dir', dir_location = ny_rev_connect_dir.split('RFPFinder')[1])
else:pass
central_hudson_dir = os.path.join(ny_rev_connect_dir, 'Central Hudson')
conedison_dir = os.path.join(ny_rev_connect_dir, 'ConEdison')
nationalgrid_dir = os.path.join(ny_rev_connect_dir, 'NationalGrid')
orange_and_rockland_dir = os.path.join(ny_rev_connect_dir, 'Orange & Rockland')
nysge_dir = os.path.join(ny_rev_connect_dir, 'NYSGE')
rge_dir = os.path.join(ny_rev_connect_dir, 'RG & E')

# San Diego Gas and Electric
san_diego_gas_and_electric_dir = os.path.join(data_dir, 'San Diego Gas and Electric')
if not os.path.exists(san_diego_gas_and_electric_dir):
    os.mkdir(san_diego_gas_and_electric_dir)
    history('created_dir', dir_location = san_diego_gas_and_electric_dir.split('RFPFinder')[1])
else:pass

###############################################################################
###############################################################################
###############################################################################

# AEP
###############################################################################

def extract_important_dates(soup):

    '''
    Parses soup html for important dates for aep sites
    Parameters:
        soup (bs4.BeautifulSoup): soup for parsing
    Returns:
        datetype_date_dict (dict): date title, date dict of all important dates
    '''
    datetype_date_dict = {}

    pars = soup.findAll('p')
    for p in pars:
        if len(p.findAll('span', attrs={'class':'bold'})) > 0:
            two_lines = p.text.replace(u'\xa0', u' ').split('\n')
            for line in [line for line in two_lines if len(line) > 0]:

                date_type, date = line.split(':  ')
                datetype_date_dict[date_type] = date

    if len(datetype_date_dict) > 0:

        # Was able to find important dates
        return datetype_date_dict

    else:

        # Keep looking
        for p in pars:
            if len(p.findAll('b')) > 0:
                dates_text = p.text
                dates_list = [s.replace('\r', '') for s in dates_text.split('\n') if len(s) > 0]
                datetype_date_dict = {s.split(': ')[0]:s.split(': ')[1] for s in dates_list}
                break
            else:pass

        return datetype_date_dict

def extract_correspondence_email(soup):

    '''
    AEP RFPs may have a correspondence email listed towards the bottom of their
    webpage, this will find those
    Parameters:
        soup (bs4.BeautifulSoup): soup for parsing
    Returns:
        correspondence_email (str): a string consisting of either an email, or None
    '''

    links = soup.findAll('a')
    try:
        correspondence_email = [link for link in [link for link in links if link.has_attr('href')] if 'mailto' in link['href'] and link.text != 'contact us'][0].text
    except:
        correspondence_email = 'None'

    return correspondence_email

def extract_rfp_desc(soup):

    '''
    Will get the first few <p> tags as long as they are incrementing by 1 in the list
    This should be redone soon...
    Parameters:
        soup (bs4.BeautifulSoup): soup for parsing
    Returns:
        new_desc (str): descripton listed on RFPs website formatted by paragraph
    '''

    content = soup.find('span', attrs={'id':'cphContentMain_GlobalUserControl1'})
    ps = [p for p in content.findAll('p')]
    new_ps = [p for p in ps if not p.find('b')]
    if len(ps) == len(new_ps):
        new_ps = [p for p in ps if not p.find('span')]
    else:pass
    shouldbe = 0
    desc = []
    for inum, p in enumerate(ps):
        if p in new_ps:
            num = inum
        else:pass
        if num != shouldbe:
            break
        else:
            desc.append(p)
        shouldbe += 1
    dstrs = [d.text for d in desc]
    new_desc = '\n \n'.join(dstrs)
    return new_desc

def aep_scrape(area_dir, url):

    '''
    Scrapes a single aep area webpage for rfps listed and saved all available information in that
    areas folder in the data dir
    Parameters:
        area_dir (var): the path to the areas folder
        url (str): URL to the areas webpage to scrape for RFPs
    Returns:

    '''

    new_rfps = []

    if not os.path.exists(area_dir):
        os.mkdir(area_dir)
        history('created_dir', dir_location = area_dir.split('RFPFinder')[1])
        history('aep', aep_type = 'New Area Folder', aep_area = area_dir.split('AEP' + os.sep)[1], aep_info = "Folder was created to store RFP folders and their information for the area")
    else:pass

    html = requests.get(url).content
    soup = BeautifulSoup(html, 'lxml')
    pars = soup.findAll('p')
    if len([p for p in pars if 'There are currently no RFPs being offered.' in p]) > 0:pass
    else:
        rfp_offers = soup.findAll('strong')

        for rfp in rfp_offers:
            try:
                a = str(rfp.find('a')['href'])
                rfp_url = str(url.split('default.aspx')[0] + a)

                rfp_html = requests.get(rfp_url).content
                rfp_soup = BeautifulSoup(rfp_html, 'lxml')

                rfp_name = rfp_soup.find('h1').text

                pdf_list_elements = rfp_soup.findAll('li', attrs={'class':'pdf'})

                # Dict of pdf title and pdf link
                pdfs_dict = {element.text:url.split('.com')[0] + '.com' + element.find('a')['href'] for element in pdf_list_elements}
                rfp_dir = os.path.join(area_dir, rfp_name)
                if not os.path.exists(rfp_dir):
                    os.mkdir(rfp_dir)
                    history('created_dir', dir_location = rfp_dir.split('RFPFinder')[1])
                    history('aep', aep_type = 'New RFP Folder', aep_area = rfp_dir.split(os.sep)[-1], aep_info = "Folder was created to store information and documents important to the RFP")
                    new_rfp = True
                else:
                    new_rfp = False

                os.chdir(rfp_dir)

                downloaded_files = os.listdir()

                if check_if_new_file('broken_pdf_files.txt'):
                    with open('broken_pdf_files.txt', 'w') as f:
                        f.write('A list of PDF file names that could not be downloaded:')
                        f.write('\n')
                        f.write('\n')
                else:pass

                broken_pdf_lines = open('broken_pdf_files.txt', 'r').read().split('\n')

                for pdf_name, pdf_url in pdfs_dict.items():
                    if '(' in pdf_name:
                        pdf_name = pdf_name.split('(')[0]
                    else:pass
                    pdf_name = pdf_name.replace(' ', '_') + '.pdf'

                    if pdf_name in broken_pdf_lines:pass
                    else:

                        try:
                            download_pdf(pdf_url, pdf_name)
                        except Exception as e:
                            with open('broken_pdf_files.txt', 'a') as f:
                                f.write(pdf_name)
                                f.write('\n')

                datetype_date_dict = extract_important_dates(rfp_soup)
                correspondence_email = extract_correspondence_email(rfp_soup)
                rfp_desc = extract_rfp_desc(rfp_soup)

                dates_string = ''
                for datetype, date in datetype_date_dict.items():
                    apstr = "     {0}: {1}".format(datetype, date) + "\n"
                    dates_string += apstr

                rfp_str = 'RFP Title: ' + rfp_name + '\n' + \
                'Correspondence Email: ' + correspondence_email + '\n' + \
                'Important Dates:' + '\n' + '\n' + \
                dates_string + '\n' + '\n' + \
                'Description:' + '\n' + \
                rfp_desc

                rfp_name_txt = rfp_name.replace(' ',  '_') + '.txt'
                with open(rfp_name_txt, 'w') as w:
                    w.write(rfp_str)
                    w.close()

                os.chdir(curr_dir)

                # Send an email out if this is a new RFP, move all emails to one place later
                if new_rfp:
                    new_rfps.append(new_rfp)
                else:pass

            except TypeError:pass

        return new_rfps

def aep():
    path_url_dict = {
        aep_ohio_dir:conf.get('aep', 'aep_ohio_url'),
        aep_texas_dir:conf.get('aep', 'aep_texas_url'),
        appalachian_power_dir:conf.get('aep', 'appalachian_power_url'),
        indiana_michigan_dir:conf.get('aep', 'indiana_michigan_url'),
        kentucky_power_dir:conf.get('aep', 'kentucky_power_url'),
        public_service_company_of_oklahoma_dir:conf.get('aep', 'public_service_company_of_oklahoma_url'),
        southwestern_electric_power_company_dir:conf.get('aep', 'southwestern_electric_power_company_url')
    }

    for area_dir, url in path_url_dict.items():
        aep_scrape(area_dir, url)

# Puerto Rico Government
###############################################################################

def puerto_rico_government(url = conf.get('puerto_rico_government', 'puerto_rico_government_url')):

    '''
    Scrapes http://www.p3.pr.gov/prepa-transformation.html for new PREPA RFPs
    Downloads PDFs from RFP links and attemps to read them
    Parameters:
        url (str): Defaults (and should not be changed) to the url above, and
                    gets that url from the config file
    Returns:
        ---
    '''

    pdf_link = 'http://www.p3.pr.gov/'

    html = requests.get(url).content
    soup = BeautifulSoup(html, 'lxml')

    rfp_links = soup.findAll('li', attrs={'class':'T-tulos-para-Comunicados-Noticias LinkStyle-Table'})
    pdf_links = [link for link in [str(pdf_link + link.find('a')['href']) for link in rfp_links] if link[-4:] == '.pdf']
    pdf_links = [link[21:] if link.startswith('http://www.p3.pr.gov/http://www.p3.pr.gov/') else link for link in pdf_links]

    if not os.path.exists(puerto_rico_government_pdf_dir):
        os.mkdir(puerto_rico_government_pdf_dir)
        history('created_dir', dir_location = puerto_rico_government_pdf_dir.split('RFPFinder')[1])
    os.chdir(puerto_rico_government_pdf_dir)

    downloaded_files = os.listdir()

    for link in pdf_links:

        try:

            fname = link.split('/assets/')[1]
            if fname not in downloaded_files:
                download_pdf(link, fname)
            else:
                # This file has already been downloaded
                pass
        except Exception as e:
            print(e)


    os.chdir(curr_dir)

# NY Rev Connect
###############################################################################

def central_hudson_scrape(url):

    html = requests.get(url).content
    soup = BeautifulSoup(html, 'lxml')

    table = soup.find('table', attrs={'width':716, 'height':390})
    table_rows = table.find_all('tr')

    rows = []
    for inum, tr in enumerate(table_rows):
        td = tr.find_all('td')
        row = [tr.text.replace('\n', '') for tr in td]

        if inum == 0:
            title_row = row
        else:
            rows.append(row)

    non_wires_alternative_opps_frame = pd.DataFrame(rows, columns=title_row)

    if check_if_new_file('Non Wires Alternative Opportunities.xlsx'):

        writer = ExcelWriter('Non Wires Alternative Opportunities.xlsx')
        non_wires_alternative_opps_frame.to_excel(writer)
        writer.save()

        history('ny_rev_connect', ny_rev_area = 'Central Hudson', ny_rev_info = 'Created Non Wire Alternative Opportunities Excel')
        history('excel_download', excel_name = 'Non Wires Alternative Opportunities.xlsx', file_path = central_hudson_dir)

    else:

        new_names = list(non_wires_alternative_opps_frame['Project Name/Description'])

        old_file_names = list(pd.read_excel('Non Wires Alternative Opportunities.xlsx')['Project Name/Description'])

        if old_file_names == new_names:pass
        else:

            writer = ExcelWriter('Non Wires Alternative Opportunities.xlsx')
            non_wires_alternative_opps_frame.to_excel(writer)
            writer.save()

            history('ny_rev_connect', ny_rev_area = 'Central Hudson', ny_rev_info = 'Updated Non Wire Alternative Opportunities Excel')


def conedison_scrape(url):

    html = requests.get(url).content
    soup = BeautifulSoup(html, 'lxml')

    table = soup.find('table', attrs={'class':'telerik-reTable-4'})

    table_head = table.findAll('th')
    table_rows = table.find_all('tr')

    for row in table_rows:

        cells = row.findAll('td')
        if len(cells) == 0:
            pass
        else:
            project_name = cells[0].text.replace('\n', '').replace('/', '')
            current_status = cells[1].text
            documents = cells[2].findAll('a')

            root_link = 'https://www.coned.com'

            docs = []
            for doc in documents:

                document_name = doc.text
                url = root_link + doc['href']

                docu = ConEdisonDocument(document_name, url)
                docs.append(docu)

            curr_proj = ConEdisonRFP(project_name, current_status, docs)

    now = datetime.now().strftime("%m/%d/%Y")

    for rfp in ConEdisonRFP.all_rfps:

        rfp_dir = os.path.join(conedison_dir, rfp.project_name)
        if not os.path.exists(rfp_dir):
            os.mkdir(rfp_dir)
            history('created_dir', dir_location = rfp_dir.split('RFPFinder')[1])
            history('ny_rev_connect', ny_rev_area = 'ConEdison', ny_rev_info = "Folder was created to store information and documents important to the RFP")
        os.chdir(rfp_dir)

        if check_if_new_file('info.txt'):

            with open('info.txt', 'w') as f:

                f.write("Project Name: {0}".format(rfp.project_name))
                f.write('\n')
                f.write("Current status as of {0}: {1}".format(now, rfp.current_status))
        else:pass

        for document in rfp.documents:
            name = document.document_name.replace(' ', '_') + '.pdf'
            try:
                download_pdf(document.url, name)
            except Exception as e:


                # Downloading excel files as pdfs right now... needs chanegd
                print(e)
        os.chdir(conedison_dir)

    return

def nationalgrid_scrape(url):

    soup_path = os.path.join(nationalgrid_dir, 'DO NOT EDIT (soup files -- test env only)')

    if not os.path.exists(soup_path):
        os.mkdir(soup_path)
        history('created_dir', dir_location = soup_path.split('RFPFinder')[1])
        history('ny_rev_connect', ny_rev_area = nationalgrid_dir.split(os.sep)[-1], ny_rev_info = "Created folder to store soup objects to compare later")
    os.chdir(soup_path)

    if check_if_new_file('DO_NOT_EDIT_THIS_FILE_first_national_grid_soup_test.txt'):

        was_change = False
        html = requests.get(url).content

        history('ny_rev_connect', ny_rev_area = nationalgrid_dir.split(os.sep)[-1], ny_rev_info = "First national grid soup file created")

        with open('DO_NOT_EDIT_THIS_FILE_first_national_grid_soup_test.txt', 'wb') as f:

            f.write(html)

    else:

        with open('DO_NOT_EDIT_THIS_FILE_first_national_grid_soup_test.txt', 'rb') as f:

            last = BeautifulSoup(f, 'lxml')

        html = requests.get(url).content

        with open('DO_NOT_EDIT_THIS_FILE_current_national_grid_soup_test.txt', 'wb') as f:

            f.write(html)

        with open('DO_NOT_EDIT_THIS_FILE_current_national_grid_soup_test.txt', 'rb') as f:

            current = BeautifulSoup(f, 'lxml')

        if last == current:

            was_change = False

            with open('DO_NOT_EDIT_THIS_FILE_first_national_grid_soup_test.txt', 'wb') as f:

                f.write(html)

        else:

            was_change = True
            change_alert = "**************************************************\n**************************************************\nTHERE WAS A CHANGE TO THE NATIONAL GRID WEBPAGE\n**************************************************\n**************************************************\n"
            print(change_alert)

    os.chdir(nationalgrid_dir)

    if was_change == True:

        now = datetime.now().strftime("%m/%d/%Y")
        history('ny_rev_connect', ny_rev_area = nationalgrid_dir.split(os.sep)[-1], ny_rev_info = "NATIONAL GRID WEBSITE UPDATED")

        with open("CHANGES_FOUND.txt", 'w') as f:

            f.write(now)
            f.write('\n\n')
            f.write(change_alert)


    return

def orange_and_rockland_scrape(url, area_dir):

    html = requests.get(url).content
    soup = BeautifulSoup(html, 'lxml')
    table = soup.findAll('table')[1]
    table_rows = table.findAll('tr')


    root_download_link = 'https://www.oru.com'

    for row in table_rows[1:]:

        cells_text = [cell.text for cell in row.findAll('td')[:-1]]

        project_name, project_type, project_size, rfp_status = cells_text

        project_name = project_name.split('\n')[0]

        project_name_dir = os.path.join(orange_and_rockland_dir, project_name)

        if os.path.exists(project_name_dir):pass
        else:
            os.mkdir(project_name_dir)
            history('created_dir', dir_location = project_name_dir.split('RFPFinder')[1])
            history('ny_rev_connect', ny_rev_area = area_dir.split(os.sep)[-1], ny_rev_info = "New RFP folder added: {0}".format(project_name))

        os.chdir(project_name_dir)

        if check_if_new_file('info.txt'):

            with open('info.txt', 'w') as f:

                f.write("Project Name: {0}".format(project_name))
                f.write('\n')
                f.write("Project Type: {0}".format(project_type))
                f.write('\n')
                f.write("Project Size: {0}".format(project_size))
                f.write('\n')
                f.write("RFP Status: {0}".format(rfp_status))
                f.write('\n')
        else:pass


        documents_cell = row.findAll('td')[-1]
        document_links = [a for a in documents_cell.findAll('a')]
        pdf_links = [a for a in document_links if '(xls)' not in a.text]
        excel_links = [a for a in document_links if '(xls)' in a.text]

        for pdf in pdf_links:

            new_url = root_download_link + pdf['href']
            pdf_text = pdf.text.replace(' ', '_') + '.pdf'
            download_pdf(new_url, pdf_text)

        for excel in excel_links:

            new_url = root_download_link + excel['href']
            excel_text = excel.text.replace(' ', '_') + '.xlsx'
            download_excel(new_url, excel_text)

        os.chdir(conedison_dir)

    return

def nysge_rge_scrape(url, area_dir):

    html = requests.get(url).content
    soup = BeautifulSoup(html, 'lxml')
    table = soup.find('table')
    table_rows = table.findAll('tr')[1:]

    for row in table_rows:

        transmission_distribution, project_name, time_of_need, estimated_release_timing = [r.text.strip() for r in row.findAll('td')]

        project_name = project_name.replace("/", '')

        project_name_dir = os.path.join(area_dir, project_name)

        if os.path.exists(project_name_dir):pass
        else:
            os.mkdir(project_name_dir)
            history('created_dir', dir_location = project_name_dir.split('RFPFinder')[1])
            history('ny_rev_connect', ny_rev_area = area_dir.split(os.sep)[-1], ny_rev_info = "New RFP folder added: {0}".format(project_name))

        os.chdir(project_name_dir)

        if check_if_new_file('info.txt'):

            with open('info.txt', 'w') as f:

                f.write("Transmission / Distribution: {0}".format(transmission_distribution))
                f.write('\n')
                f.write("Project Name: {0}".format(project_name))
                f.write('\n')
                f.write("Time of Need: {0}".format(time_of_need))
                f.write('\n')
                f.write("Estimated Release Timing: {0}".format(estimated_release_timing))
                f.write('\n')
        else:pass

        os.chdir(area_dir)

    return

def ny_rev_connect_scrape(area_dir, url):

    if not os.path.exists(area_dir):
        os.mkdir(area_dir)
        history('created_dir', dir_location = area_dir.split('RFPFinder')[1])
        history('ny_rev_connect', ny_rev_area = area_dir.split(os.sep)[-1], ny_rev_info = "Folder was created to store RFP folders and their information for the area")
    os.chdir(area_dir)

    if area_dir == central_hudson_dir:
        central_hudson_scrape(url)
    elif area_dir == conedison_dir:
        conedison_scrape(url)
    elif area_dir == nationalgrid_dir:
        nationalgrid_scrape(url)
    elif area_dir == orange_and_rockland_dir:
        orange_and_rockland_scrape(url, area_dir)
    elif area_dir in [nysge_dir, rge_dir]:
        nysge_rge_scrape(url, area_dir)

    os.chdir(curr_dir)
    return

def ny_rev_connect():
    path_url_dict = {
        central_hudson_dir:conf.get('ny_rev_connect', 'central_hudson_url'),
        conedison_dir:conf.get('ny_rev_connect', 'conedison_url'),
        nationalgrid_dir:conf.get('ny_rev_connect', 'nationalgrid_url'),
        orange_and_rockland_dir:conf.get('ny_rev_connect', 'orange_and_rockland_url'),
        nysge_dir:conf.get('ny_rev_connect', 'nysge_url'),
        rge_dir:conf.get('ny_rev_connect', 'rge_url'),
    }

    for area_dir, url in path_url_dict.items():
        ny_rev_connect_scrape(area_dir, url)

    return

#
###############################################################################

def get_sdge_rfp_info(rfp, open_closed_dir):

    try:

        node = rfp['href']
        rfp_url = 'https://www.sdge.com{0}'.format(node)
        rfp_html = requests.get(rfp_url).content
        rfp_soup = BeautifulSoup(rfp_html, 'lxml')

        rfp_title = rfp_soup.find('h1', attrs={'class':"builder-title col-sm-12"}).text.replace('/', '')

        sdge_rfp_area_dir = os.path.join(open_closed_dir, rfp_title)
        if not os.path.exists(sdge_rfp_area_dir):
            os.mkdir(sdge_rfp_area_dir)
            history('created_dir', dir_location = sdge_rfp_area_dir.split('RFPFinder')[1])
        os.chdir(sdge_rfp_area_dir)




        # DOWNLOADING DOCUMENTS
        sdge_rfp_area_docs_dir = os.path.join(sdge_rfp_area_dir, 'Documents')
        if not os.path.exists(sdge_rfp_area_docs_dir):
            os.mkdir(sdge_rfp_area_docs_dir)
            history('created_dir', dir_location = sdge_rfp_area_docs_dir.split('RFPFinder')[1])
        os.chdir(sdge_rfp_area_docs_dir)

        for a in rfp_soup.findAll('a'):
            try:
                if a['href'].endswith('.pdf'):

                    pdf_name = a.text + '.pdf'

                    print(a['href'])

                    download_pdf(a['href'], pdf_name)

            except KeyError:
                pass

        os.chdir(sdge_rfp_area_dir)

        content_div = rfp_soup.find('div', attrs={'class':'builder-element-inside-inner'})

        title_pars = {}
        for h2 in content_div.findAll('h2'):

            ps = h2.find_all_next('p')
            title_pars[h2] = ps

        real_title_pars = {}
        found_real_pars = []
        for title, pars in title_pars.items():

            other_pars = [par_list for h2, par_list in title_pars.items() if h2 != title]
            op = [item for sublist in other_pars for item in sublist]

            real_pars = []
            for par in pars:

                if par not in op:

                    real_pars.append(par)
                    found_real_pars.append(par)

            real_title_pars[title] = real_pars
            title_pars[title] = real_pars

        info_txt = ''

        for title, par in real_title_pars.items():

            if len(par) > 1:

                par = ''.join([x.text for x in par])

            else:

                par = par[0].text

            info_txt += "{}\n{}\n\n".format(title.text, par)

        with open('info.txt', 'w') as f:

            f.write(info_txt)



        for t, pl in title_pars.items():

            if t.text == 'RFP Schedule':

                # table = rfp_soup.find('table', attrs={'width':'587'})
                # table_rows = table.find_all('tr')
                #
                # rows = []
                # for inum, tr in enumerate(table_rows):
                #     td = tr.find_all('td')
                #     row = [tr.text.replace('\n', '') for tr in td]
                #
                #     if inum == 0:
                #         title_row = row
                #     else:
                #         rows.append(row)
                #
                # non_wires_alternative_opps_frame = pd.DataFrame(rows, columns=title_row)
                #
                # if check_if_new_file('RFP Schedule.xlsx'):
                #
                #     writer = ExcelWriter('RFP Schedule.xlsx')
                #     non_wires_alternative_opps_frame.to_excel(writer)
                #     writer.save()
                #
                #     history('excel_download', excel_name = 'RFP Schedule.xlsx', file_path = sdge_rfp_area_dir)

                pass

            else:
                pass

    except Exception as e:

        print(e)

    finally:

        os.chdir(open_closed_dir)

    return

def san_diego_gas_and_electric_scrape():

    os.chdir(san_diego_gas_and_electric_dir)

    url = conf.get('san_diego_gas_and_electric', 'san_diego_gas_and_electric_url')

    html = requests.get(url).content
    soup = BeautifulSoup(html, 'lxml')

    closed_rfps = [h for h in soup.findAll('h2') if h.text == 'Closed RFO/RFP'][0]
    closed_rfps_list = closed_rfps.find_all_next('a')

    open_rfps = [h for h in soup.findAll('h1') if h.text == 'RFPs and RFOs'][0]
    open_rfps_list = [a for a in open_rfps.find_all_next('a') if a not in closed_rfps_list]

    open_rfps_dir = os.path.join(san_diego_gas_and_electric_dir, 'Open RFPS')
    if not os.path.exists(open_rfps_dir):
        os.mkdir(open_rfps_dir)
        history('created_dir', dir_location = san_diego_gas_and_electric_dir.split('RFPFinder')[1])
    os.chdir(open_rfps_dir)

    for rfp_link_obj in open_rfps_list:

        get_sdge_rfp_info(rfp_link_obj, open_rfps_dir)

    os.chdir(san_diego_gas_and_electric_dir)

    closed_rfps_dir = os.path.join(san_diego_gas_and_electric_dir, 'Closed RFPS')
    if not os.path.exists(closed_rfps_dir):
        os.mkdir(closed_rfps_dir)
        history('created_dir', dir_location = san_diego_gas_and_electric_dir.split('RFPFinder')[1])
    os.chdir(closed_rfps_dir)

    for rfp_link_obj in closed_rfps_list:

        get_sdge_rfp_info(rfp_link_obj, closed_rfps_dir)

    os.chdir(san_diego_gas_and_electric_dir)
    os.chdir(curr_dir)

    return


# Main
###############################################################################
def main():

    # aep()
    # print('   - AEP finished')
    # puerto_rico_government()
    # print('   - Government of Puerto Rico finished')
    # ny_rev_connect()
    # print('   - NY Rev Connect finished')
    san_diego_gas_and_electric_scrape()
    print('   - San Diego Gas and Electric finished')

###############################################################################
###############################################################################
###############################################################################

if __name__ == '__main__':

    main()
